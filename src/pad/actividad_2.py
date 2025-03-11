import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns  # type: ignore
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

class Ejercicio:
    def __init__(self):
        self.ruta_excel = r"/workspaces/Act_1-Maria-ospina/src/pad/resultado.xlsx"
        self.resultados = []  # Lista para guardar los resultados

    def guardar_grafico(self, filename):
        """Guarda el gráfico actual y lo cierra para liberar memoria"""
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        plt.close()

    def resolver_puntos(self):
        """Resuelve los puntos y guarda los resultados"""

        # 🔹 Punto 1
        array_10_29 = np.arange(10, 30)
        self.resultados.append(["Punto 1", "Array 10-29", str(array_10_29.tolist())])

        # 🔹 Punto 2
        suma_total = np.ones((10, 10)).sum()
        self.resultados.append(["Punto 2", "Suma Matriz 10x10", suma_total])

        # 🔹 Punto 3
        array1 = np.random.randint(1, 11, size=5)
        array2 = np.random.randint(1, 11, size=5)
        producto_elemento = array1 * array2
        self.resultados.append(["Punto 3", "Array 1", str(array1.tolist())])
        self.resultados.append(["Punto 3", "Array 2", str(array2.tolist())])
        self.resultados.append(["Punto 3", "Producto Elemento a Elemento", str(producto_elemento.tolist())])

        # 🔹 Punto 4
        i = np.arange(4).reshape(4, 1)
        j = np.arange(4).reshape(1, 4)
        matriz = i + j
        try:
            inv_matriz = np.linalg.inv(matriz)
        except np.linalg.LinAlgError:
            inv_matriz = np.linalg.pinv(matriz)

        self.resultados.append(["Punto 4", "Matriz 4x4 (i+j)", str(matriz.tolist())])
        self.resultados.append(["Punto 4", "Matriz Inversa o Pseudo-inversa", str(inv_matriz.tolist())])

        # 🔹 Punto 5
        arr_aleatorio = np.random.rand(100)
        self.resultados.append(["Punto 5", "Valor Máximo", arr_aleatorio.max()])
        self.resultados.append(["Punto 5", "Valor Mínimo", arr_aleatorio.min()])

        # 🔹 Punto 6
        array_3x1 = np.array([[1], [2], [3]])
        array_1x3 = np.array([[10, 20, 30]])
        self.resultados.append(["Punto 6", "Suma con Broadcasting", str((array_3x1 + array_1x3).tolist())])

        # 🔹 Punto 7
        matriz_5x5 = np.arange(1, 26).reshape(5, 5)
        self.resultados.append(["Punto 7", "Submatriz 2x2", str(matriz_5x5[1:3, 1:3].tolist())])

        # 🔹 Punto 8
        array_ceros = np.zeros(10)
        array_ceros[3:7] = 5
        self.resultados.append(["Punto 8", "Array de ceros modificado", str(array_ceros.tolist())])

        # 🔹 Punto 9
        matriz_3x3 = np.arange(1, 10).reshape(3, 3)
        self.resultados.append(["Punto 9", "Matriz con filas invertidas", str(matriz_3x3[::-1, :].tolist())])

        # 🔹 Punto 10
        array_random = np.random.rand(10)
        self.resultados.append(["Punto 10", "Elementos > 0.5", str(array_random[array_random > 0.5].tolist())])

        # 🔹 Punto 11: Gráfico de dispersión
        plt.figure()
        plt.scatter(np.random.rand(100), np.random.rand(100), color='blue', alpha=0.5)
        plt.title("Punto 11 - Gráfico de Dispersión")
        self.guardar_grafico("grafico_dispersion.png")

        # 🔹 Punto 12: Gráfico de dispersión con y = sin(x) + ruido
        x = np.linspace(-2*np.pi, 2*np.pi, 100)
        y = np.sin(x) + np.random.normal(0, 0.1, size=x.shape)
        plt.figure()
        plt.scatter(x, y, color='red', alpha=0.5)
        plt.plot(x, np.sin(x), color='black', label="$y = sin(x)$")
        plt.legend()
        plt.title("Punto 12 - Dispersión con y=sin(x) + ruido")
        self.guardar_grafico("grafico_seno.png")

        # 🔹 Punto 13: Gráfico de contorno
        X, Y = np.meshgrid(np.linspace(-5, 5, 100), np.linspace(-5, 5, 100))
        Z = np.cos(X) + np.sin(Y)
        plt.figure()
        plt.contour(X, Y, Z, levels=20, cmap='viridis')
        plt.colorbar()
        plt.title("Punto 13 - Gráfico de Contorno")
        self.guardar_grafico("grafico_contorno.png")

        # 🔹 Punto 14: Gráfico de dispersión con densidad
        plt.figure()
        sns.kdeplot(x=np.random.rand(100), y=np.random.rand(100), fill=True, cmap='Blues')
        plt.title("Punto 14 - Dispersión con Densidad")
        self.guardar_grafico("grafico_densidad.png")

        # 🔹 Punto 15: Gráfico de contorno lleno
        plt.figure()
        plt.contourf(X, Y, Z, levels=20, cmap='plasma')
        plt.colorbar()
        plt.title("Punto 15 - Contorno Lleno")
        self.guardar_grafico("grafico_contorno_lleno.png")

        # 🔹 Punto 16: Gráfico de dispersión con leyenda
        plt.figure()
        plt.scatter(x, y, color='red', label="$y = sin(x) + ruido$", alpha=0.5)
        plt.plot(x, np.sin(x), color='black', label="$y = sin(x)$")
        plt.xlabel("Eje X")
        plt.ylabel("Eje Y")
        plt.legend()
        plt.title("Punto 16 - Dispersión con Leyenda")
        self.guardar_grafico("grafico_sin_ruido.png")

        # 🔹 Guardar resultados en Excel
        df = pd.DataFrame(self.resultados, columns=["Punto", "Descripción", "Resultado"])
        df.to_excel(self.ruta_excel, index=False)

        # 🔹 Insertar gráficos en Excel con títulos
        wb = load_workbook(self.ruta_excel)
        ws = wb.active

        fila_imagen = len(self.resultados) + 6

        for punto, archivo in [
            ("Punto 11 - Gráfico de Dispersión", "grafico_dispersion.png"),
            ("Punto 12 - Dispersión con y=sin(x) + ruido", "grafico_seno.png"),
            ("Punto 13 - Gráfico de Contorno", "grafico_contorno.png"),
            ("Punto 14 - Dispersión con Densidad", "grafico_densidad.png"),
            ("Punto 15 - Contorno Lleno", "grafico_contorno_lleno.png"),
            ("Punto 16 - Dispersión con Leyenda", "grafico_sin_ruido.png")
        ]:
            ws[f"A{fila_imagen}"] = punto
            img = Image(archivo)
            img.width, img.height = 400, 300
            ws.add_image(img, f"B{fila_imagen + 1}")
            fila_imagen += 17

        wb.save(self.ruta_excel)
        print(f"Los resultados y gráficos han sido guardados en '{self.ruta_excel}'.")

# 🔹 Ejecutar la clase
ej = Ejercicio()
ej.resolver_puntos()

import openpyxl

# Cargar el archivo Excel
ruta_archivo = "/workspaces/Act_1-Maria-ospina/src/pad/resultado.xlsx"
wb = openpyxl.load_workbook(ruta_archivo)

# Seleccionar la primera hoja
hoja = wb.active

# Leer algunas celdas (por ejemplo, las primeras 5 filas y 3 columnas)
for fila in hoja.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3, values_only=True):
    print(fila)

# Cerrar el archivo
wb.close()
